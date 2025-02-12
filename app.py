"""Streamlit web application for Jira sprint analysis."""

import streamlit as st
import pandas as pd
import os
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

from jira_analysis.analyzer import process_discipline_data
from jira_analysis.visualization import create_comparison_plot, format_excel_cell
from jira_analysis.utils import ensure_directory

st.set_page_config(page_title="Jira Sprint Analysis", layout="wide")

st.title("Jira Sprint Analysis Dashboard")

# File uploader
uploaded_file = st.file_uploader("Upload your Jira Excel file", type=['xlsx'])

if uploaded_file is not None:
    # Create input directory if it doesn't exist
    ensure_directory("input")
    
    # Save the uploaded file
    input_path = os.path.join("input", uploaded_file.name)
    with open(input_path, "wb") as f:
        f.write(uploaded_file.getvalue())
else:
    # If no file is uploaded, try to use existing file in input directory
    input_files = [f for f in os.listdir("input") if f.endswith('.xlsx')]
    if input_files:
        input_path = os.path.join("input", input_files[0])
    else:
        st.error("No Excel file found. Please upload a file.")
        st.stop()

try:
    # Read the Excel file
    df = pd.read_excel(input_path, skiprows=3)
    
    # Filter out rows with blank or NaN keys
    df = df[df["Key"].notna() & (df["Key"] != "")]
    
    # Display total unique tickets
    st.metric("Total Unique Tickets", len(df["Key"].unique()))
    
    if st.button("Process Analysis"):
        with st.spinner("Processing data..."):
            # Ensure output directory exists
            output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")
            temp_dir = os.path.join(output_dir, "temp")
            ensure_directory(output_dir)
            ensure_directory(temp_dir)
            
            # Define disciplines
            disciplines = ["QA", "TA", "FE", "BE", "BA"]
            
            # Column mapping (same as in original script)
            column_mapping = {
                "QA": {
                    "original": "QA Original Estimate based on old way",
                    "ai": "QA Revised Estimate based on AI",
                    "actual": "QA Actual Time based on AI",
                },
                "TA": {
                    "original": "TA Original Estimate based on old way",
                    "ai": "TA Revised Estimate based on AI",
                    "actual": "TA Actual Time based on AI",
                },
                "FE": {
                    "original": "FE Dev Original Estimate based on old way",
                    "ai": "FE Dev Revised Estimate based on AI",
                    "actual": "FE Dev Actual Time based on AI",
                },
                "BE": {
                    "original": "BE Dev Original Estimate based on old way",
                    "ai": "BE Dev Revised Estimate based on AI",
                    "actual": "BE Dev Actual Time based on AI",
                },
                "BA": {
                    "original": "BA Original Estimate based on old way",
                    "ai": "BA Revised Estimate based on AI",
                    "actual": "BA Actual Time based on AI",
                },
            }
            
            # Calculate total unique tickets
            total_unique_tickets = len(df["Key"].unique())

            # Process data for each discipline
            analysis_results = []
            all_missing_data = []  # Collect missing data from all disciplines
            
            for discipline, mapping in column_mapping.items():
                discipline_data = process_discipline_data(df, mapping)
                
                # Add to totals
                total_original_estimate = discipline_data["Original Estimate (Total)"]
                total_ai_estimate = discipline_data["AI Estimate (Total)"]
                total_actual_time = discipline_data["Actual Time (Total)"]
                total_complete = discipline_data["Complete Data Points"]
                total_partial = discipline_data["Partial Data Points"]
                total_missing = discipline_data["Missing Data Points"]

                analysis_results.append({
                    "Discipline": discipline,
                    **discipline_data
                })
                
                # Add discipline to missing data details
                for detail in discipline_data["Missing Data Details"]:
                    detail["Discipline"] = discipline
                    all_missing_data.append(detail)

            # Create analysis DataFrame
            analysis_df = pd.DataFrame(analysis_results)
            
            # Display results
            st.subheader("Estimation Analysis by Discipline")
            
            # Calculate percentage of actual vs original
            analysis_df["Actual vs Original (%)"] = (analysis_df["Actual Time (Total)"] / analysis_df["Original Estimate (Total)"] * 100).round(1)
            
            try:
                # Create visualization first (using numeric data)
                fig = create_comparison_plot(analysis_df, disciplines)
                
                # Display plot in Streamlit
                st.subheader("Visualization")
                st.pyplot(fig)
                
                # Now format the data for display
                display_df = analysis_df.copy()
                
                # Format percentages
                display_df["Actual vs Original (%)"] = display_df["Actual vs Original (%)"].fillna(0).apply(lambda x: f"{x}%")
                
                # Format the numeric columns with 1 decimal place and add 'h' suffix
                for col in ["Original Estimate (Total)", "AI Estimate (Total)", "Actual Time (Total)"]:
                    display_df[col] = display_df[col].round(1).apply(lambda x: f"{x}h")
                
                # Select and rename columns for display
                display_df = display_df[[
                    "Discipline",
                    "Original Estimate (Total)",
                    "AI Estimate (Total)",
                    "Actual Time (Total)",
                    "Partial Data Points",
                    "Actual vs Original (%)"
                ]].rename(columns={
                    "Original Estimate (Total)": "Original",
                    "AI Estimate (Total)": "AI Estimate",
                    "Actual Time (Total)": "Actual",
                    "Partial Data Points": "Missing Data Points",
                    "Actual vs Original (%)": "Actual vs Original"
                })
                
                # Display the table
                st.table(display_df)
                
                # Add explanation
                st.write("""
                **Metrics Explanation:**
                - **Original**: Total hours originally estimated
                - **AI Estimate**: Total hours estimated by AI
                - **Actual**: Total hours actually spent
                - **Missing Data Points**: Number of tickets missing either AI estimates or actual time
                - **Actual vs Original**: Percentage of actual time compared to original estimate (e.g., 60% means actual took 60% of original estimate)
                """)
                
                # Display missing data details if any exist
                if all_missing_data:
                    st.subheader("Tickets with Incomplete Data")
                    missing_df = pd.DataFrame(all_missing_data)
                    
                    # Format the values as a readable string
                    def format_values(row):
                        values = []
                        for field in ["Original Estimate", "AI Estimate", "Actual Time"]:
                            if field in row["Values"] and row["Values"][field] != 0:
                                values.append(f"{field}: {row['Values'][field]:.1f}h")
                        return "\n".join(values)
                    
                    missing_df["Present Values"] = missing_df.apply(format_values, axis=1)
                    missing_df["Missing Fields"] = missing_df["Missing Fields"].apply(lambda x: ", ".join(x) if isinstance(x, list) else '')
                    
                    # Sort by discipline and ticket
                    missing_df = missing_df.sort_values(["Discipline", "Ticket"])
                    
                    # Display with custom column order
                    st.table(missing_df[["Discipline", "Ticket", "Assignee", "Present Values", "Missing Fields"]])
                    
                    st.write("""
                    This table shows tickets that have at least one non-zero value but are missing others:
                    - Each ticket shows the values that are present (excluding zeros)
                    - Missing fields include both null values and zeros
                    - Fields can be: Original Estimate, AI Estimate, or Actual Time
                    """)
                    
                    # Group by missing field combinations
                    st.subheader("Missing Fields Summary")
                    missing_combos = missing_df.groupby("Missing Fields").size().reset_index(name="Count")
                    missing_combos = missing_combos.sort_values("Count", ascending=False)
                    st.table(missing_combos)
                    
                # Display Learnings by Discipline last
                st.subheader("Learnings by Discipline")
                
                # Create tabs for each discipline
                disciplines = ["TA", "BA", "QA", "BE Dev", "FE Dev"]
                tabs = st.tabs(disciplines)
                
                for tab, discipline in zip(tabs, disciplines):
                    with tab:
                        # Filter for the specific discipline's learnings
                        discipline_col = f"{discipline} Learnings"
                        if discipline_col in df.columns:
                            discipline_data = df[["Key", "Summary", discipline_col]].copy()
                            
                            def is_valid_learning(x):
                                if pd.isna(x):
                                    return False
                                # Convert to string and lowercase
                                x = str(x).lower().strip()
                                # Check if empty or just whitespace
                                if not x:
                                    return False
                                # Check against invalid values
                                invalid_values = {'0', '0.0', 'none', 'n/a', 'nan'}
                                if x in invalid_values:
                                    return False
                                # Check if it's any variation of zero (00.00, 0.0, etc)
                                if x.replace('.', '').replace('0', '') == '':
                                    return False
                                return True
                            
                            # Apply the filtering function
                            discipline_data = discipline_data[discipline_data[discipline_col].apply(is_valid_learning)]
                            
                            if not discipline_data.empty:
                                # Convert dataframe to a format suitable for st.table
                                table_data = discipline_data.copy()
                                # Rename the discipline column to "Learnings"
                                table_data.columns = ["Key", "Summary", "Learnings"]
                                st.table(table_data)
                            else:
                                st.info(f"No learnings found for {discipline}")
                
                # Save plot for Excel
                img_bytes = BytesIO()
                fig.savefig(img_bytes, format='png', bbox_inches="tight", dpi=300)
                img_bytes.seek(0)
                
                # Create Excel report
                wb = Workbook()
                
                # Analysis Results sheet
                ws_analysis = wb.active
                ws_analysis.title = "Analysis Results"
                
                # Add title
                format_excel_cell(ws_analysis, "A1", "Jira Sprint Analysis Report", bold=True)
                format_excel_cell(ws_analysis, "A2", f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
                
                # Convert display_df to rows, replacing any empty lists with None
                excel_df = display_df.copy()
                excel_rows = [
                    ['' if isinstance(val, list) and len(val) == 0 else val 
                     for val in row]
                    for row in dataframe_to_rows(excel_df, index=False, header=True)
                ]
                
                # Add the rows to the worksheet
                for r_idx, row in enumerate(excel_rows, 4):
                    for c_idx, value in enumerate(row, 1):
                        ws_analysis.cell(row=r_idx, column=c_idx, value=value)
                
                # Create Missing Data sheet if there are missing data points
                if all_missing_data:
                    ws_missing = wb.create_sheet("Missing Data")
                    
                    # Add title
                    format_excel_cell(ws_missing, "A1", "Tickets with Incomplete Data", bold=True)
                    format_excel_cell(ws_missing, "A2", "Shows tickets that have some fields populated but are missing others")
                    
                    # Add missing data
                    missing_df = pd.DataFrame(all_missing_data)
                    
                    # Format the values as a readable string
                    def format_values(row):
                        values = []
                        for field in ["Original Estimate", "AI Estimate", "Actual Time"]:
                            if field in row["Values"] and row["Values"][field] != 0:
                                values.append(f"{field}: {row['Values'][field]:.1f}h")
                        return "\n".join(values)
                    
                    missing_df["Present Values"] = missing_df.apply(format_values, axis=1)
                    missing_df["Missing Fields"] = missing_df["Missing Fields"].apply(lambda x: ", ".join(x) if isinstance(x, list) else '')
                    missing_df = missing_df.sort_values(["Discipline", "Ticket"])
                    
                    # Add headers with formatting
                    headers = ["Discipline", "Ticket", "Assignee", "Present Values", "Missing Fields"]
                    for col, header in enumerate(headers, 1):
                        format_excel_cell(ws_missing, f"{chr(64+col)}4", header, bold=True)
                    
                    # Convert missing_df to rows, replacing any empty lists with empty strings
                    missing_rows = [
                        ['' if isinstance(val, list) and len(val) == 0 else val 
                         for val in row]
                        for row in dataframe_to_rows(missing_df[headers], index=False, header=False)
                    ]
                    
                    # Add data
                    for r_idx, row in enumerate(missing_rows, 5):
                        for c_idx, value in enumerate(row, 1):
                            cell = ws_missing.cell(row=r_idx, column=c_idx, value=value)
                            # Enable text wrapping for the Present Values column
                            if c_idx == 4:  # Present Values column
                                cell.alignment = Alignment(wrapText=True)
                    
                    # Adjust column widths
                    for col in ws_missing.columns:
                        max_length = 0
                        for cell in col:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)  # Cap width at 50 characters
                        ws_missing.column_dimensions[col[0].column_letter].width = adjusted_width
                    
                    # Add row height for better readability of wrapped text
                    for row in ws_missing.iter_rows(min_row=5):
                        ws_missing.row_dimensions[row[0].row].height = 45
                    
                    # Add Missing Fields Summary sheet
                    ws_summary = wb.create_sheet("Missing Fields Summary")
                    format_excel_cell(ws_summary, "A1", "Missing Fields Summary", bold=True)
                    format_excel_cell(ws_summary, "A2", "Count of tickets by missing field combinations")
                    
                    # Create summary table
                    missing_combos = missing_df.groupby("Missing Fields").size().reset_index(name="Count")
                    missing_combos = missing_combos.sort_values("Count", ascending=False)
                    
                    # Add headers
                    format_excel_cell(ws_summary, "A4", "Missing Fields", bold=True)
                    format_excel_cell(ws_summary, "B4", "Count", bold=True)
                    
                    # Add data
                    for r_idx, row in enumerate(dataframe_to_rows(missing_combos, index=False, header=False), 5):
                        for c_idx, value in enumerate(row, 1):
                            ws_summary.cell(row=r_idx, column=c_idx, value=value)
                    
                    # Adjust column widths
                    ws_summary.column_dimensions['A'].width = 40
                    ws_summary.column_dimensions['B'].width = 10
                
                # Add Learnings sheets for each discipline
                disciplines_with_learnings = ["QA", "TA", "FE Dev", "BE Dev", "BA"]
                for discipline in disciplines_with_learnings:
                    discipline_col = f"{discipline} Learnings"
                    if discipline_col in df.columns:
                        discipline_data = df[["Key", "Summary", discipline_col]].copy()
                        
                        # Apply the same filtering as in the display
                        discipline_data = discipline_data[discipline_data[discipline_col].apply(is_valid_learning)]
                        
                        if not discipline_data.empty:
                            # Create sheet for this discipline
                            ws_discipline = wb.create_sheet(f"{discipline} Learnings")
                            
                            # Add title and explanation
                            format_excel_cell(ws_discipline, "A1", f"{discipline} Learnings", bold=True)
                            format_excel_cell(ws_discipline, "A2", f"Learnings from {discipline} tickets")
                            
                            # Rename the column for clarity
                            discipline_data = discipline_data.rename(columns={discipline_col: "Learnings"})
                            
                            # Add headers
                            headers = ["Key", "Summary", "Learnings"]
                            for col, header in enumerate(headers, 1):
                                format_excel_cell(ws_discipline, f"{chr(64+col)}4", header, bold=True)
                            
                            # Add data
                            for r_idx, row in enumerate(dataframe_to_rows(discipline_data, index=False, header=False), 5):
                                for c_idx, value in enumerate(row, 1):
                                    cell = ws_discipline.cell(row=r_idx, column=c_idx, value=value)
                                    # Enable text wrapping for Summary and Learnings
                                    if c_idx in [2, 3]:  # Summary and Learnings columns
                                        cell.alignment = Alignment(wrapText=True)
                            
                            # Adjust column widths
                            ws_discipline.column_dimensions['A'].width = 15  # Key
                            ws_discipline.column_dimensions['B'].width = 50  # Summary
                            ws_discipline.column_dimensions['C'].width = 50  # Learnings
                            
                            # Add row height for better readability of wrapped text
                            for row in ws_discipline.iter_rows(min_row=5):
                                ws_discipline.row_dimensions[row[0].row].height = 45
                
                # Save Excel file
                excel_path = os.path.join(output_dir, f"analysis_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
                wb.save(excel_path)
                
                # Add visualization as a new sheet
                wb = load_workbook(excel_path)
                ws_viz = wb.create_sheet("Visualization")
                img = Image(img_bytes)
                ws_viz.add_image(img, 'A1')
                
                # Resave with visualization
                wb.save(excel_path)
                
                # Provide download link
                with open(excel_path, "rb") as f:
                    bytes_data = f.read()
                    st.download_button(
                        label="Download Excel Report",
                        data=bytes_data,
                        file_name=os.path.basename(excel_path),
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )                
            except Exception as e:
                st.error(f"Error creating visualization or Excel report: {str(e)}")
                
except Exception as e:
    st.error(f"An error occurred while processing the data: {str(e)}")
