import os
import glob
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def find_input_file():
    """Find the most recent Excel file in the input directory."""
    files = glob.glob("input/*.xlsx")
    if not files:
        raise FileNotFoundError(
            "No Excel files found in the input directory. "
            "Please place your Jira export file in the input folder."
        )
    # Get the most recently modified file
    latest_file = max(files, key=os.path.getmtime)
    print(f"Using input file: {latest_file}")
    return latest_file

def ensure_output_dir():
    """Ensure the output directory exists."""
    os.makedirs("output", exist_ok=True)

# Find and read the input Excel file
try:
    input_file = find_input_file()
    df = pd.read_excel(input_file, skiprows=3)
except Exception as e:
    print(f"Error reading input file: {e}")
    exit(1)

# Filter out rows with blank or NaN keys
df = df[df["Key"].notna() & (df["Key"] != "")]

# Define the mapping for each discipline's columns
column_mapping = {
    "QA": {
        "original": "QA Original Estimate based on old way",
        "ai": "QA Revised Estimate based on AI",
        "actual": "QA Actual Time based on AI",
    },
    "TA": {
        "original": "TA Original Estimate based on old way",
        "ai": "TA Revised Estimate based on AI",
        "actual": "TA Actual Time based on AI",
    },
    "FE": {
        "original": "FE Dev Original Estimate based on old way",
        "ai": "FE Dev Revised Estimate based on AI",
        "actual": "FE Dev Actual Time based on AI",
    },
    "BE": {
        "original": "BE Dev Original Estimate based on old way",
        "ai": "BE Dev Revised Estimate based on AI",
        "actual": "BE Dev Actual Time based on AI",
    },
    "BA": {
        "original": "BA Original Estimate based on old way",
        "ai": "BA Revised Estimate based on AI",
        "actual": "BA Actual Time based on AI",
    },
}

# List of disciplines
disciplines = list(column_mapping.keys())

# Create analysis dataframe
analysis_results = []

# Create missing data analysis
missing_data_results = []


def format_cell(ws, cell, value, color=None, bold=False, alignment=None):
    """Format a cell in the worksheet.

    Args:
        ws: Worksheet to format
        cell: Cell reference
        value: Value to set
        color: Fill color (optional)
        bold: Whether to make text bold (optional)
        alignment: Cell alignment (optional)
    """
    ws[cell] = value
    if color:
        ws[cell].fill = PatternFill(
            start_color=color, end_color=color, fill_type="solid"
        )
    if bold:
        ws[cell].font = Font(bold=True)
    if alignment:
        ws[cell].alignment = alignment


def calculate_percentage_diff(original, actual):
    """Calculate percentage difference between original and actual values.

    Args:
        original: Original value
        actual: Actual value

    Returns:
        float: Percentage difference
    """
    if original == 0:
        return 0
    return ((original - actual) / original) * 100


def process_discipline_data(df, discipline_mapping):
    """Process data for a specific discipline.

    Args:
        df: DataFrame containing the data
        discipline_mapping: Dictionary containing column mappings

    Returns:
        dict: Dictionary containing the processed data
    """
    result = {
        "Complete Data Points": 0,
        "Missing Data Points": 0,
        "Original Estimate (Total)": 0,
        "AI Estimate (Total)": 0,
        "Actual Time (Total)": 0,
    }

    for _, row in df.iterrows():
        original_col = discipline_mapping["original"]
        ai_col = discipline_mapping["ai"]
        actual_col = discipline_mapping["actual"]

        if pd.notna(row[original_col]):
            if pd.notna(row[ai_col]) and pd.notna(row[actual_col]):
                result["Complete Data Points"] += 1
                result["Original Estimate (Total)"] += float(row[original_col])
                result["AI Estimate (Total)"] += float(row[ai_col])
                result["Actual Time (Total)"] += float(row[actual_col])
            else:
                result["Missing Data Points"] += 1
                result["Original Estimate (Total)"] += float(row[original_col])

    return result


# Process each row to find missing data
for idx, row in df.iterrows():
    for discipline in disciplines:
        original_col = column_mapping[discipline]["original"]
        ai_col = column_mapping[discipline]["ai"]
        actual_col = column_mapping[discipline]["actual"]

        if pd.notna(row[original_col]):
            missing_fields = []
            if pd.isna(row[ai_col]):
                missing_fields.append("AI Estimate")
            if pd.isna(row[actual_col]):
                missing_fields.append("Actual Time")

            if missing_fields:
                missing_data_results.append(
                    {
                        "Key": row["Key"],
                        "Discipline": discipline,
                        "Missing Fields": ", ".join(missing_fields),
                    }
                )

# Create missing data DataFrame and sort by Discipline
missing_df = pd.DataFrame(missing_data_results)
if not missing_df.empty:
    missing_df = missing_df.sort_values(["Discipline", "Key"])

# Process complete data for analysis
total_original_estimate = 0
total_ai_estimate = 0
total_actual_time = 0
total_complete_tickets = 0
total_incomplete_tickets = 0

# Process data for each discipline
for discipline in disciplines:
    discipline_data = process_discipline_data(df, column_mapping[discipline])

    # Update totals
    total_original_estimate += discipline_data["Original Estimate (Total)"]
    total_ai_estimate += discipline_data["AI Estimate (Total)"]
    total_actual_time += discipline_data["Actual Time (Total)"]
    total_complete_tickets += discipline_data["Complete Data Points"]
    total_incomplete_tickets += discipline_data["Missing Data Points"]

    # Calculate differences
    diff_original = calculate_percentage_diff(
        discipline_data["Original Estimate (Total)"],
        discipline_data["Actual Time (Total)"],
    )
    diff_ai = calculate_percentage_diff(
        discipline_data["AI Estimate (Total)"], discipline_data["Actual Time (Total)"]
    )

    analysis_results.append(
        {
            "Discipline": discipline,
            "Complete Data Points": discipline_data["Complete Data Points"],
            "Missing Data Points": discipline_data["Missing Data Points"],
            "Original Estimate (Total)": round(
                discipline_data["Original Estimate (Total)"], 2
            ),
            "AI Estimate (Total)": round(discipline_data["AI Estimate (Total)"], 2),
            "Actual Time (Total)": round(discipline_data["Actual Time (Total)"], 2),
            "Diff from Original (%)": round(diff_original, 2),
            "Diff from AI (%)": round(diff_ai, 2),
        }
    )

analysis_df = pd.DataFrame(analysis_results)

# Calculate overall completion rate
overall_completion_rate = (
    (total_complete_tickets / (total_complete_tickets + total_incomplete_tickets) * 100)
    if (total_complete_tickets + total_incomplete_tickets) > 0
    else 0
)

# Only proceed if we have any results
if len(analysis_df) > 0:
    # Function to add value labels on bars
    def add_value_labels(ax, bars):
        for bar in bars:
            height = bar.get_height()
            ax.text(
                bar.get_x() + bar.get_width() / 2.0,
                height,
                f"{height:.2f}",
                ha="center",
                va="bottom",
            )

    # Create visualizations
    fig, ax = plt.subplots(figsize=(12, 6))
    x = range(len(analysis_df))
    width = 0.25

    bars1 = ax.bar(
        [i - width for i in x],
        analysis_df["Original Estimate (Total)"],
        width,
        label="Original Estimate",
        color="blue",
    )
    bars2 = ax.bar(
        x, analysis_df["AI Estimate (Total)"], width, label="AI Estimate", color="green"
    )
    bars3 = ax.bar(
        [i + width for i in x],
        analysis_df["Actual Time (Total)"],
        width,
        label="Actual Time",
        color="red",
    )

    # Add value labels
    add_value_labels(ax, bars1)
    add_value_labels(ax, bars2)
    add_value_labels(ax, bars3)

    plt.xlabel("Disciplines")
    plt.ylabel("Total Hours")
    plt.title(
        "Comparison of Total Estimates vs Actual Time\n(Only Complete Data Points)"
    )
    plt.xticks(
        x,
        analysis_df["Discipline"]
        + "\n(n="
        + analysis_df["Complete Data Points"].astype(str)
        + ")",
    )
    plt.legend()

    temp_plot_file = "estimate_comparison.png"
    try:
        # Try to save the plot as PNG
        plt.savefig(temp_plot_file, bbox_inches="tight")
    except (ValueError, OSError, RuntimeError) as e:
        print(f"Error saving plot: {e}")

    plt.close()

    # Create percentage difference plot
    fig, ax = plt.subplots(figsize=(12, 6))
    x = range(len(analysis_df))
    width = 0.35

    bars1 = ax.bar(
        [i - width / 2 for i in x],
        analysis_df["Diff from Original (%)"],
        width,
        label="Difference from Original (%)",
        color="purple",
    )
    bars2 = ax.bar(
        [i + width / 2 for i in x],
        analysis_df["Diff from AI (%)"],
        width,
        label="Difference from AI (%)",
        color="orange",
    )

    # Add value labels
    add_value_labels(ax, bars1)
    add_value_labels(ax, bars2)

    plt.xlabel("Disciplines")
    plt.ylabel("Difference (%)")
    plt.title("Percentage Difference from Estimates\n(Only Complete Data Points)")
    plt.xticks(
        x,
        analysis_df["Discipline"]
        + "\n(n="
        + analysis_df["Complete Data Points"].astype(str)
        + ")",
    )
    plt.legend()

    temp_plot_file = "percentage_difference.png"
    try:
        # Try to save the plot as PNG
        plt.savefig(temp_plot_file, bbox_inches="tight")
    except (ValueError, OSError, RuntimeError) as e:
        print(f"Error saving plot: {e}")

    plt.close()

    # Create Excel workbook with results
    wb = Workbook()

    # Summary Dashboard
    ws_summary = wb.active
    ws_summary.title = "Summary Dashboard"

    # Define styles
    header_font = Font(bold=True, size=12)
    subheader_font = Font(bold=True)
    normal_font = Font(size=11)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Add title
    ws_summary["A1"] = "Estimation Analysis Summary"
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary.merge_cells("A1:E1")
    ws_summary["A1"].alignment = Alignment(horizontal="center")

    # Overall Statistics
    ws_summary["A3"] = "Overall Statistics"
    ws_summary["A3"].font = header_font

    # Add overall statistics
    stats_data = [
        ("Total Complete Tickets:", total_complete_tickets),
        ("Total Incomplete Tickets:", total_incomplete_tickets),
        ("Overall Completion Rate:", f"{overall_completion_rate:.1f}%"),
        ("Total Original Estimate (Hours):", f"{total_original_estimate:.2f}"),
        ("Total AI Estimate (Hours):", f"{total_ai_estimate:.2f}"),
        ("Total Actual Time (Hours):", f"{total_actual_time:.2f}"),
    ]

    for idx, (label, value) in enumerate(stats_data):
        row = idx + 4
        ws_summary[f"A{row}"] = label
        ws_summary[f"B{row}"] = value
        ws_summary[f"A{row}"].font = normal_font
        ws_summary[f"B{row}"].font = normal_font

    # Add completion rates by discipline
    ws_summary["A11"] = "Discipline Analysis"
    ws_summary["A11"].font = header_font

    headers = [
        "Discipline",
        "Complete",
        "Incomplete",
        "Completion Rate",
        "Original Est. (Total)",
        "AI Est. (Total)",
        "Actual (Total)",
        "AI vs Original Diff %",
    ]
    for col, header in enumerate(headers):
        ws_summary.cell(row=12, column=col + 1, value=header).font = subheader_font

    for idx, row in analysis_df.iterrows():
        data_row = 13 + idx
        complete = row["Complete Data Points"]
        incomplete = row["Missing Data Points"]
        completion_rate = (
            (complete / (complete + incomplete) * 100)
            if (complete + incomplete) > 0
            else 0
        )

        # Calculate AI vs Original difference
        original_est = row["Original Estimate (Total)"]
        ai_est = row["AI Estimate (Total)"]
        ai_vs_original = (
            ((original_est - ai_est) / original_est * 100) if original_est != 0 else 0
        )

        ws_summary.cell(row=data_row, column=1, value=row["Discipline"])
        ws_summary.cell(row=data_row, column=2, value=complete)
        ws_summary.cell(row=data_row, column=3, value=incomplete)
        ws_summary.cell(row=data_row, column=4, value=f"{completion_rate:.1f}%")
        ws_summary.cell(row=data_row, column=5, value=original_est).number_format = (
            "0.00"
        )
        ws_summary.cell(row=data_row, column=6, value=ai_est).number_format = "0.00"
        ws_summary.cell(
            row=data_row, column=7, value=row["Actual Time (Total)"]
        ).number_format = "0.00"

        # Add conditional formatting for AI vs Original difference
        diff_cell = ws_summary.cell(
            row=data_row, column=8, value=f"{ai_vs_original:.1f}%"
        )
        if ai_vs_original > 0:  # AI estimate is lower than original
            diff_cell.fill = PatternFill(
                start_color="90EE90", end_color="90EE90", fill_type="solid"
            )  # Light green
        else:  # AI estimate is higher than original
            diff_cell.fill = PatternFill(
                start_color="FFB6C1", end_color="FFB6C1", fill_type="solid"
            )  # Light red

    # Adjust column widths
    for col in range(1, ws_summary.max_column + 1):
        max_length = 0
        for row in range(1, ws_summary.max_row + 1):
            cell = ws_summary.cell(row=row, column=col)
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2
        ws_summary.column_dimensions[get_column_letter(col)].width = adjusted_width

    # Analysis Results Sheet
    ws_analysis = wb.create_sheet("Analysis Results")

    # Write the analysis dataframe
    for r_idx, r in enumerate(dataframe_to_rows(analysis_df, index=False, header=True)):
        ws_analysis.append(r)

        # Format numbers to 2 decimal places
        if r_idx > 0:  # Skip header row
            for col in range(
                4, 9
            ):  # Columns D to H (all numeric columns except Complete/Missing Data Points)
                cell = ws_analysis.cell(row=r_idx + 1, column=col)
                cell.number_format = "0.00"

            # Format Complete/Missing Data Points as whole numbers
            cell = ws_analysis.cell(row=r_idx + 1, column=2)  # Complete Data Points
            cell.number_format = "0"
            cell = ws_analysis.cell(row=r_idx + 1, column=3)  # Missing Data Points
            cell.number_format = "0"

        # Add conditional formatting to the percentage columns
        if r_idx > 0:  # Skip header row
            # Get the cells for percentage differences
            orig_diff_cell = ws_analysis.cell(row=r_idx + 1, column=7)  # Column G
            ai_diff_cell = ws_analysis.cell(row=r_idx + 1, column=8)  # Column H

            # Define colors
            green_fill = PatternFill(
                start_color="90EE90", end_color="90EE90", fill_type="solid"
            )  # Light green
            red_fill = PatternFill(
                start_color="FFB6C1", end_color="FFB6C1", fill_type="solid"
            )  # Light red

            # Apply formatting for Original estimate difference
            if orig_diff_cell.value > 0:
                orig_diff_cell.fill = green_fill
            else:
                orig_diff_cell.fill = red_fill

            # Apply formatting for AI estimate difference
            if ai_diff_cell.value > 0:
                ai_diff_cell.fill = green_fill
            else:
                ai_diff_cell.fill = red_fill

    # Missing Data Sheet
    if len(missing_df) > 0:
        ws_missing = wb.create_sheet("Missing Data")

        # Write the missing data dataframe
        for r_idx, r in enumerate(
            dataframe_to_rows(missing_df, index=False, header=True)
        ):
            ws_missing.append(r)

            # Highlight entire row in light yellow
            if r_idx > 0:  # Skip header row
                yellow_fill = PatternFill(
                    start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
                )
                for cell in ws_missing[r_idx + 1]:
                    cell.fill = yellow_fill

                # Format Original Estimate to 2 decimal places
                cell = ws_missing.cell(
                    row=r_idx + 1, column=4
                )  # Original Estimate column
                cell.number_format = "0.00"

        # Adjust column widths for better readability
        for column in ws_missing.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max_length + 2
            ws_missing.column_dimensions[get_column_letter(column[0].column)].width = (
                adjusted_width
            )

    # Original Data Sheet
    ws_original = wb.create_sheet("Original Data")

    # Write the original dataframe
    for r_idx, r in enumerate(dataframe_to_rows(df, index=False, header=True)):
        ws_original.append(r)

    # Adjust column widths for better readability
    for column in ws_original.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(
            (max_length + 2), 50
        )  # Cap width at 50 to prevent too wide columns
        ws_original.column_dimensions[get_column_letter(column[0].column)].width = (
            adjusted_width
        )

    # Visualizations Sheet
    ws2 = wb.create_sheet("Visualizations")
    col_offset = 0
    for temp_plot_file in ["estimate_comparison.png", "percentage_difference.png"]:
        try:
            # Try to add the plot to the Excel file
            img = Image(temp_plot_file)
            cell_ref = f"{get_column_letter(col_offset + 1)}1"
            ws2.add_image(img, cell_ref)
        except (ValueError, OSError, RuntimeError) as e:
            print(f"Error adding plot to Excel: {e}")
        col_offset += 6

    ensure_output_dir()
    output_file = "output/estimation_analysis.xlsx"
    try:
        # Try to create the Excel file
        wb.save(output_file)
        print(f"Analysis complete! Check '{output_file}' for results.")
    except (PermissionError, OSError) as e:
        print(f"Error saving output file: {e}")
        exit(1)

    try:
        # Try to clean up temporary files
        for temp_file in ["estimate_comparison.png", "percentage_difference.png"]:
            try:
                os.remove(temp_file)
            except (OSError, FileNotFoundError) as e:
                print(f"Warning: Could not remove temporary file {temp_file}: {e}")
    except Exception as e:
        print(f"Error during cleanup: {e}")
else:
    print("No complete data points found for any discipline.")
