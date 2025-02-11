"""Visualization utilities for Jira sprint analysis."""

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from typing import List, Tuple, Any
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.worksheet import Worksheet

def create_comparison_plot(
    analysis_df: pd.DataFrame,
    disciplines: List[str],
    figsize: Tuple[int, int] = (8, 5)
) -> plt.Figure:
    """Create a bar plot comparing estimates and actuals.
    
    Args:
        analysis_df: DataFrame containing the analysis results
        disciplines: List of discipline names
        figsize: Figure size tuple (width, height)
    
    Returns:
        Figure object
    """
    fig, ax = plt.subplots(figsize=figsize)
    x = np.arange(len(disciplines))
    width = 0.25

    original_estimates = [analysis_df.loc[analysis_df['Discipline'] == d, 'Original Estimate (Total)'].values[0] for d in disciplines]
    ai_estimates = [analysis_df.loc[analysis_df['Discipline'] == d, 'AI Estimate (Total)'].values[0] for d in disciplines]
    actual_times = [analysis_df.loc[analysis_df['Discipline'] == d, 'Actual Time (Total)'].values[0] for d in disciplines]

    rects1 = ax.bar(x - width, original_estimates, width, label='Original Estimate', color='skyblue')
    rects2 = ax.bar(x, ai_estimates, width, label='AI Estimate', color='lightgreen')
    rects3 = ax.bar(x + width, actual_times, width, label='Actual Time', color='salmon')

    ax.set_ylabel('Hours')
    ax.set_title('Comparison of Estimates vs Actual Time by Discipline')
    ax.set_xticks(x)
    ax.set_xticklabels(disciplines)
    ax.legend()

    def autolabel(rects):
        for rect in rects:
            height = rect.get_height()
            ax.annotate(f'{height:.1f}',
                       xy=(rect.get_x() + rect.get_width() / 2, height),
                       xytext=(0, 3),  # 3 points vertical offset
                       textcoords="offset points",
                       ha='center', va='bottom', rotation=0)

    autolabel(rects1)
    autolabel(rects2)
    autolabel(rects3)

    fig.tight_layout()
    return fig

def format_excel_cell(
    ws: Worksheet,
    cell: str,
    value: Any,
    color: str = None,
    bold: bool = False,
    alignment: str = None
) -> None:
    """Format a cell in the Excel worksheet.
    
    Args:
        ws: Worksheet to format
        cell: Cell reference
        value: Value to set
        color: Fill color (optional)
        bold: Whether to make text bold (optional)
        alignment: Cell alignment (optional)
    """
    ws[cell].value = value
    if color:
        ws[cell].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    if bold:
        ws[cell].font = Font(bold=True)
    if alignment:
        ws[cell].alignment = Alignment(horizontal=alignment)
    
    # Add border
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    ws[cell].border = border
