"""Tests for the visualization module."""

import unittest
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from jira_analysis.visualization import create_comparison_plot, format_excel_cell

class TestVisualization(unittest.TestCase):
    def setUp(self):
        """Set up test data."""
        self.analysis_df = pd.DataFrame({
            'Discipline': ['QA', 'TA'],
            'Original Estimate (Total)': [5.0, 6.0],
            'AI Estimate (Total)': [4.0, 5.0],
            'Actual Time (Total)': [4.6, 5.5],
        })
        self.disciplines = ['QA', 'TA']

    def test_create_comparison_plot(self):
        """Test plot creation."""
        fig = create_comparison_plot(self.analysis_df, self.disciplines)
        self.assertIsInstance(fig, plt.Figure)
        plt.close(fig)  # Clean up

    def test_format_excel_cell(self):
        """Test Excel cell formatting."""
        wb = Workbook()
        ws = wb.active
        
        # Test basic value setting
        format_excel_cell(ws, "A1", "Test")
        self.assertEqual(ws["A1"].value, "Test")
        
        # Test bold formatting
        format_excel_cell(ws, "A2", "Bold", bold=True)
        self.assertTrue(ws["A2"].font.bold)
        
        # Test color formatting
        format_excel_cell(ws, "A3", "Colored", color="FFFF0000")
        self.assertEqual(ws["A3"].fill.start_color.rgb, "FFFF0000")
        
        # Test alignment
        format_excel_cell(ws, "A4", "Centered", alignment="center")
        self.assertEqual(ws["A4"].alignment.horizontal, "center")

if __name__ == "__main__":
    unittest.main()
